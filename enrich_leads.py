#!/usr/bin/env python3
"""
Shepherd's Guild — Detail Enrichment Pass
==========================================
Reads the locate_church_raw.csv from the initial scrape, filters to
multi-directory churches (listed on 2+ of 9Marks/Founders/TGC), fetches
each church's detail page for pastor/email/phone/website, and writes
enriched outputs.

Usage:
  python3 enrich_leads.py                          # Enrich multi-directory only (default)
  python3 enrich_leads.py --all                     # Enrich ALL churches (slow)
  python3 enrich_leads.py --input path/to/raw.csv   # Custom input path
  python3 enrich_leads.py --min-score 1             # Enrich any church with score >= 1
"""

import argparse
import csv
import json
import os
import re
import random
import time
import logging
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}
DELAY_MIN = 1.0
DELAY_MAX = 2.5
MAX_RETRIES = 3

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("enricher")

session = requests.Session()
session.headers.update(HEADERS)

PROGRESS_FILE = "enrich_progress.json"


# ============================================================
# FETCH + PARSE
# ============================================================
def fetch_detail(url):
    for attempt in range(MAX_RETRIES):
        try:
            time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
            resp = session.get(url, timeout=30)
            if resp.status_code == 200:
                return resp.text
            elif resp.status_code in (403, 429):
                wait = 10 * (attempt + 1)
                log.warning(f"HTTP {resp.status_code}, waiting {wait}s — {url}")
                time.sleep(wait)
            else:
                log.warning(f"HTTP {resp.status_code}: {url}")
                return None
        except requests.RequestException as e:
            log.error(f"Error: {e} (attempt {attempt+1})")
            time.sleep(5)
    return None


def parse_detail(html):
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(separator="\n", strip=True)
    data = {"pastor": "", "email": "", "phone": "", "address": "", "website": ""}

    # Pastor
    m = re.search(r"Pastor:\s*(.+?)(?:\n|$)", text)
    if m:
        data["pastor"] = m.group(1).strip()

    # Email
    email_link = soup.find("a", href=re.compile(r"^mailto:"))
    if email_link:
        data["email"] = email_link["href"].replace("mailto:", "").strip()
    else:
        m = re.search(r"Email:\s*(\S+@\S+)", text)
        if m:
            data["email"] = m.group(1).strip()

    # Phone
    m = re.search(r"Phone:\s*([\d\-\(\)\.\s]+)", text)
    if m:
        data["phone"] = m.group(1).strip()

    # Website
    for a in soup.find_all("a", href=re.compile(r"^https?://")):
        href = a["href"]
        if "locate.church" not in href and "thegospelcoalition" not in href:
            data["website"] = href
            break

    # Address
    m = re.search(
        r"(\d+\s+[\w\s\.]+(?:Street|St|Avenue|Ave|Road|Rd|Drive|Dr|Boulevard|Blvd|Lane|Ln|Way|Circle|Ct|Court|Pike|Highway|Hwy)[\w\s\.]*)",
        text, re.IGNORECASE,
    )
    if m:
        data["address"] = m.group(1).strip()

    return data


# ============================================================
# PROGRESS
# ============================================================
def save_progress(enriched, remaining_slugs):
    with open(PROGRESS_FILE, "w") as f:
        json.dump({
            "enriched": enriched,
            "remaining": remaining_slugs,
            "timestamp": datetime.now().isoformat(),
        }, f)


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return None


# ============================================================
# READ CSV
# ============================================================
def read_csv(filepath):
    churches = []
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            churches.append(row)
    return churches


def is_multi_directory(row):
    flags = [
        row.get("is_9marks", "").strip().lower() == "yes",
        row.get("is_founders", "").strip().lower() == "yes",
        row.get("is_tgc", "").strip().lower() == "yes",
    ]
    return sum(flags) >= 2


def directory_count(row):
    flags = [
        row.get("is_9marks", "").strip().lower() == "yes",
        row.get("is_founders", "").strip().lower() == "yes",
        row.get("is_tgc", "").strip().lower() == "yes",
    ]
    return sum(flags)


# ============================================================
# XLSX OUTPUT
# ============================================================
def write_enriched_xlsx(churches, filepath):
    HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="Arial", size=10, color="333333")
    GOLD_FILL = PatternFill("solid", fgColor="FFF8E1")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Church Leads"
    ws.sheet_properties.tabColor = "1A1A2E"

    headers = [
        "Lead ID", "Church Name", "Senior Pastor", "Pastor Email",
        "Church Email", "Phone", "Website", "City", "State", "ZIP",
        "Denomination", "Est. Congregation Size", "Directory Source(s)",
        "9Marks Listed", "Founders Listed", "TGC Listed",
        "Fiscal Year Model", "Sermon Archive Online?", "Archive Platform",
        "Est. Sermon Count", "Lead Status", "Lead Score",
        "First Contact Date", "Last Contact Date", "Contact Method",
        "Notes", "Next Action", "Next Action Date", "Assigned To",
        "Source URL",
    ]

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    widths = [10, 30, 22, 28, 28, 16, 32, 16, 8, 10, 20, 14, 22,
              12, 12, 10, 16, 14, 18, 14, 16, 10, 14, 14, 14, 32, 24, 14, 14, 36]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Sort: multi-directory first, then by directory count desc
    churches.sort(key=lambda c: directory_count(c), reverse=True)

    for idx, church in enumerate(churches):
        row = idx + 2
        lead_id = f"SG-{idx+1:04d}"

        dc = directory_count(church)
        score = 3 if dc >= 2 else (1 if dc == 1 else 0)

        is_9m = church.get("is_9marks", "No")
        is_fn = church.get("is_founders", "No")
        is_tgc = church.get("is_tgc", "No")

        row_data = [
            lead_id,
            church.get("name", ""),
            church.get("pastor", ""),
            church.get("email", ""),
            "",  # Church Email
            church.get("phone", ""),
            church.get("website", ""),
            church.get("city", ""),
            church.get("state", ""),
            church.get("zip", ""),
            "",  # Denomination
            "",  # Est. Congregation Size
            church.get("networks", ""),
            is_9m,
            is_fn,
            is_tgc,
            "",  # Fiscal Year Model
            "",  # Sermon Archive Online?
            "",  # Archive Platform
            "",  # Est. Sermon Count
            "Research" if dc >= 2 else "New",
            score if score > 0 else "",
            "",  # First Contact Date
            "",  # Last Contact Date
            "",  # Contact Method
            church.get("description", ""),
            "Verify pastor email" if church.get("email") else "Research contact info",
            "",  # Next Action Date
            "",  # Assigned To
            church.get("detail_url", ""),
        ]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            # Highlight multi-directory rows
            if dc >= 2:
                cell.fill = GOLD_FILL

    max_row = max(len(churches) + 1, 2)

    dv_status = DataValidation(type="list", formula1='"New,Research,Contacted,Warm,Qualified,Proposal Sent,Won,Lost,Nurture"')
    ws.add_data_validation(dv_status)
    dv_status.add(f"U2:U{max_row}")

    dv_fiscal = DataValidation(type="list", formula1='"Calendar (Jan-Dec),Ministry (Sep-Aug),Unknown,Other"')
    ws.add_data_validation(dv_fiscal)
    dv_fiscal.add(f"Q2:Q{max_row}")

    for col_range in ["N", "O", "P", "R"]:
        dv = DataValidation(type="list", formula1='"Yes,No,Unknown"')
        ws.add_data_validation(dv)
        dv.add(f"{col_range}2:{col_range}{max_row}")

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Enrichment Summary")
    ws2.sheet_properties.tabColor = "C5A55A"

    enriched_count = sum(1 for c in churches if c.get("pastor") or c.get("email"))
    multi_count = sum(1 for c in churches if directory_count(c) >= 2)

    summary = [
        ["Shepherd's Guild — Enriched Lead List"],
        [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        [""],
        ["Metric", "Value"],
        ["Total churches", len(churches)],
        ["Multi-directory (2+)", multi_count],
        ["With pastor name", sum(1 for c in churches if c.get("pastor"))],
        ["With email", sum(1 for c in churches if c.get("email"))],
        ["With phone", sum(1 for c in churches if c.get("phone"))],
        ["With website", sum(1 for c in churches if c.get("website"))],
        ["Enrichment rate", f"{enriched_count}/{len(churches)} ({enriched_count*100//max(len(churches),1)}%)"],
    ]
    for r, row_data in enumerate(summary, 1):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val).font = Font(name="Arial", size=10)
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    ws2.merge_cells("A1:B1")
    ws2["A4"].font = Font(name="Arial", bold=True, size=10)
    ws2["B4"].font = Font(name="Arial", bold=True, size=10)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    wb.save(filepath)
    log.info(f"XLSX written: {filepath} ({len(churches)} rows, {multi_count} multi-directory)")


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Enrich Locate.Church leads with detail page data")
    parser.add_argument("--input", default="locate_church_raw.csv", help="Path to raw CSV from initial scrape")
    parser.add_argument("--all", action="store_true", help="Enrich ALL churches, not just multi-directory")
    parser.add_argument("--min-score", type=int, default=None, help="Enrich churches with directory count >= this")
    parser.add_argument("--resume", action="store_true", help="Resume from progress file")
    parser.add_argument("--output-dir", default=".", help="Output directory")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)

    # Read the raw CSV
    if not os.path.exists(args.input):
        log.error(f"Input file not found: {args.input}")
        log.info("Run the initial scrape first: python3 scrape_locate_church.py --skip-details")
        sys.exit(1)

    import sys
    churches = read_csv(args.input)
    log.info(f"Loaded {len(churches)} churches from {args.input}")

    # Determine which churches to enrich
    if args.all:
        to_enrich = churches
        log.info(f"Enriching ALL {len(to_enrich)} churches")
    elif args.min_score is not None:
        to_enrich = [c for c in churches if directory_count(c) >= args.min_score]
        log.info(f"Enriching {len(to_enrich)} churches with {args.min_score}+ directory listings")
    else:
        to_enrich = [c for c in churches if is_multi_directory(c)]
        log.info(f"Enriching {len(to_enrich)} multi-directory churches")

    not_enriching = [c for c in churches if c not in to_enrich]

    # Resume support
    already_done = {}
    if args.resume:
        progress = load_progress()
        if progress and progress.get("enriched"):
            already_done = {e["slug"]: e for e in progress["enriched"]}
            log.info(f"Resumed with {len(already_done)} already-enriched churches")

    # Enrich
    enriched = []
    total = len(to_enrich)
    for i, church in enumerate(to_enrich):
        slug = church.get("slug", "")

        # Skip if already done (resume)
        if slug in already_done:
            church.update(already_done[slug])
            enriched.append(church)
            continue

        detail_url = church.get("detail_url", "")
        if not detail_url:
            enriched.append(church)
            continue

        if i % 25 == 0:
            log.info(f"Progress: {i}/{total} ({i*100//max(total,1)}%)")

        html = fetch_detail(detail_url)
        if html:
            detail = parse_detail(html)
            church["pastor"] = detail.get("pastor", "")
            church["email"] = detail.get("email", "")
            church["phone"] = detail.get("phone", "")
            church["address"] = detail.get("address", "")
            if detail.get("website"):
                church["website"] = detail["website"]

        enriched.append(church)

        # Save progress every 50 churches
        if i > 0 and i % 50 == 0:
            remaining = [c.get("slug", "") for c in to_enrich[i+1:]]
            save_progress(enriched, remaining)

    log.info(f"Enrichment complete: {len(enriched)} churches processed")

    # Combine enriched + non-enriched
    all_churches = enriched + not_enriching

    # Write outputs
    csv_path = str(output_dir / "sg_enriched_leads.csv")
    xlsx_path = str(output_dir / "sg-church-crm-enriched.xlsx")

    # Write enriched CSV
    fieldnames = list(all_churches[0].keys()) if all_churches else []
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_churches)
    log.info(f"CSV written: {csv_path}")

    # Write enriched XLSX
    write_enriched_xlsx(all_churches, xlsx_path)

    # Summary
    with_pastor = sum(1 for c in enriched if c.get("pastor"))
    with_email = sum(1 for c in enriched if c.get("email"))
    with_phone = sum(1 for c in enriched if c.get("phone"))
    multi = sum(1 for c in enriched if is_multi_directory(c))

    print("\n" + "=" * 60)
    print("ENRICHMENT COMPLETE")
    print("=" * 60)
    print(f"Churches enriched: {len(enriched)}")
    print(f"  Multi-directory: {multi}")
    print(f"  With pastor name: {with_pastor}")
    print(f"  With email: {with_email}")
    print(f"  With phone: {with_phone}")
    print(f"Total churches in output: {len(all_churches)}")
    print(f"\nOutputs:")
    print(f"  CSV:  {csv_path}")
    print(f"  XLSX: {xlsx_path}")
    print("=" * 60)

    # Cleanup
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)


if __name__ == "__main__":
    main()
