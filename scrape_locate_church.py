#!/usr/bin/env python3
"""
Shepherd's Guild — Locate.Church Directory Scraper
===================================================
Scrapes church listings from locate.church for 9Marks, Founders Ministries,
and TGC networks. Two-pass approach:
  Pass 1: Collect church stubs from paginated list pages (name, location, networks, detail URL)
  Pass 2: Visit each detail page for enriched data (pastor, email, phone, address, website)

Outputs to the SG Church CRM spreadsheet format.

Usage:
  python scrape_locate_church.py                    # Full scrape, all networks, all pages
  python scrape_locate_church.py --max-pages 3      # Limit to 3 pages per network (test run)
  python scrape_locate_church.py --skip-details      # List pages only, no detail enrichment
  python scrape_locate_church.py --networks 9marks   # Single network only
  python scrape_locate_church.py --resume             # Resume from existing progress file
"""

import argparse
import csv
import json
import os
import re
import sys
import time
import random
import logging
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

# ============================================================
# CONFIG
# ============================================================
BASE_URL = "https://locate.church"
NETWORKS = {
    "9marks": "/networks/9marks",
    "founders": "/networks/founders-ministries",
    "tgc": "/networks/the-gospel-coalition",
}
# Fallback URL patterns (locate.church uses both /networks/ and /denominations/)
NETWORK_FALLBACKS = {
    "9marks": ["/denominations/9marks", "/networks/9-marks"],
    "founders": ["/denominations/founders-ministries", "/networks/founders"],
    "tgc": ["/denominations/the-gospel-coalition", "/networks/gospel-coalition"],
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

DELAY_MIN = 1.5  # seconds between requests
DELAY_MAX = 3.5
DETAIL_DELAY_MIN = 1.0
DETAIL_DELAY_MAX = 2.5
MAX_RETRIES = 3

PROGRESS_FILE = "scrape_progress.json"
OUTPUT_CSV = "locate_church_raw.csv"
OUTPUT_XLSX = "sg-church-crm-populated.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("scraper")


# ============================================================
# HTTP
# ============================================================
session = requests.Session()
session.headers.update(HEADERS)


def fetch_page(url, retries=MAX_RETRIES):
    """Fetch a page with retry logic and polite delays."""
    for attempt in range(retries):
        try:
            delay = random.uniform(DELAY_MIN, DELAY_MAX)
            time.sleep(delay)
            resp = session.get(url, timeout=30)
            if resp.status_code == 200:
                return resp.text
            elif resp.status_code == 403:
                log.warning(f"403 Forbidden: {url} (attempt {attempt+1})")
                time.sleep(5 * (attempt + 1))
            elif resp.status_code == 429:
                wait = 30 * (attempt + 1)
                log.warning(f"Rate limited on {url}, waiting {wait}s")
                time.sleep(wait)
            else:
                log.warning(f"HTTP {resp.status_code}: {url}")
                time.sleep(3)
        except requests.RequestException as e:
            log.error(f"Request error: {e} (attempt {attempt+1})")
            time.sleep(5 * (attempt + 1))
    log.error(f"Failed after {retries} attempts: {url}")
    return None


def fetch_detail(url):
    """Fetch a detail page with shorter delays."""
    for attempt in range(MAX_RETRIES):
        try:
            delay = random.uniform(DETAIL_DELAY_MIN, DETAIL_DELAY_MAX)
            time.sleep(delay)
            resp = session.get(url, timeout=30)
            if resp.status_code == 200:
                return resp.text
            elif resp.status_code in (403, 429):
                wait = 10 * (attempt + 1)
                log.warning(f"HTTP {resp.status_code} on detail page, waiting {wait}s")
                time.sleep(wait)
            else:
                log.warning(f"HTTP {resp.status_code}: {url}")
                return None
        except requests.RequestException as e:
            log.error(f"Detail fetch error: {e}")
            time.sleep(5)
    return None


# ============================================================
# PARSING — LIST PAGES
# ============================================================
def parse_list_page(html, network_name):
    """Parse a paginated list page and return church stubs + next page URL."""
    soup = BeautifulSoup(html, "html.parser")
    churches = []

    # Church cards are typically h2 > a links with /churches/ in href
    # Also look for divs/articles containing church info
    church_links = []

    # Method 1: Find all links to /churches/ pages
    for a_tag in soup.find_all("a", href=True):
        href = a_tag.get("href", "")
        if "/churches/" in href and "View Details" not in a_tag.get_text():
            church_links.append(a_tag)

    # Build church entries by finding name links, then scanning siblings for context
    seen_slugs = set()

    # Collect all /churches/ links with their surrounding context
    all_a_tags = soup.find_all("a", href=True)
    church_anchors = []
    for a_tag in all_a_tags:
        href = a_tag.get("href", "")
        if "/churches/" not in href:
            continue
        text = a_tag.get_text(strip=True)
        if not text or text in ("View Details", "LOCATE.CHURCH", ""):
            continue
        if len(text) < 3:
            continue
        slug = href.rstrip("/").split("/")[-1]
        if slug in seen_slugs or not slug:
            continue
        seen_slugs.add(slug)
        church_anchors.append((a_tag, slug, text, href))

    for a_tag, slug, text, href in church_anchors:
        detail_url = urljoin(BASE_URL, href)

        # Walk up to the nearest container that holds this card's info
        parent = a_tag.parent
        card_text = ""
        for _ in range(8):
            if parent is None:
                break
            sibs_text = parent.get_text(separator="\n", strip=True)
            # A good card container has location OR networks but is not the entire page
            if ("📍" in sibs_text or "Networks:" in sibs_text) and len(sibs_text) < 1000:
                card_text = sibs_text
                break
            parent = parent.parent

        if not card_text:
            # Fallback: grab text from parent up to a reasonable depth
            parent = a_tag.parent
            for _ in range(4):
                if parent:
                    card_text = parent.get_text(separator="\n", strip=True)
                    parent = parent.parent
                    if len(card_text) > 30:
                        break

        # Parse location from THIS card's text
        location = ""
        loc_match = re.search(r"📍\s*(.+?)(?:\n|$)", card_text)
        if loc_match:
            location = loc_match.group(1).strip()

        # Parse networks from THIS card's text
        networks_list = []
        net_match = re.search(r"Networks?:\s*(.+?)(?:\n|$)", card_text)
        if net_match:
            raw = net_match.group(1).strip()
            networks_list = [n.strip() for n in raw.split(",") if n.strip()]

        # Parse description
        description = ""
        lines = [l.strip() for l in card_text.split("\n") if l.strip()]
        for line in lines:
            if (line != text and "📍" not in line and "Networks" not in line
                    and "View Details" not in line and len(line) > 20
                    and not line.startswith("Showing")):
                description = line[:200]
                break

        churches.append({
            "name": text,
            "slug": slug,
            "detail_url": detail_url,
            "location_raw": location,
            "networks": networks_list if networks_list else [network_name],
            "description": description,
        })

    # Find next page URL
    next_url = None
    next_link = soup.find("a", string=re.compile(r"Next"))
    if next_link and next_link.get("href"):
        next_url = urljoin(BASE_URL, next_link["href"])

    # Also check for numbered pagination
    if not next_url:
        for a in soup.find_all("a", href=True):
            if "page=" in a.get("href", ""):
                # We'll handle this in the main loop
                pass

    # Get total count if shown
    total = None
    count_match = re.search(r"of\s+([\d,]+)\s+churches", html)
    if count_match:
        total = int(count_match.group(1).replace(",", ""))

    return churches, next_url, total


def parse_location(location_raw):
    """Parse 'City, ST' or 'City ST ZIP' or 'City ST ZIP, ST' into components."""
    city, state, zipcode = "", "", ""
    if not location_raw:
        return city, state, zipcode

    raw = location_raw.strip().rstrip(",")

    # Try pattern: "City ST ZIP"
    m = re.match(r"^(.+?)\s+([A-Z]{2})\s+(\d{5})", raw)
    if m:
        city, state, zipcode = m.group(1).rstrip(","), m.group(2), m.group(3)
        return city, state, zipcode

    # Try pattern: "City, ST"
    m = re.match(r"^(.+?),?\s+([A-Z]{2})$", raw)
    if m:
        city, state = m.group(1).rstrip(","), m.group(2)
        return city, state, zipcode

    # Try pattern with extra state at end: "City ST ZIP, ST"
    m = re.match(r"^(.+?)\s+([A-Z]{2})\s+(\d{5}),?\s*[A-Z]{2}$", raw)
    if m:
        city, state, zipcode = m.group(1).rstrip(","), m.group(2), m.group(3)
        return city, state, zipcode

    # Fallback: just return the raw string as city
    city = raw
    return city, state, zipcode


# ============================================================
# PARSING — DETAIL PAGES
# ============================================================
def parse_detail_page(html):
    """Parse a church detail page for pastor, email, phone, address, website."""
    soup = BeautifulSoup(html, "html.parser")
    data = {
        "pastor": "",
        "email": "",
        "phone": "",
        "address": "",
        "website": "",
    }

    text = soup.get_text(separator="\n", strip=True)

    # Pastor
    pastor_match = re.search(r"Pastor:\s*(.+?)(?:\n|$)", text)
    if pastor_match:
        data["pastor"] = pastor_match.group(1).strip()

    # Email — from mailto links
    email_link = soup.find("a", href=re.compile(r"^mailto:"))
    if email_link:
        data["email"] = email_link["href"].replace("mailto:", "").strip()
    else:
        email_match = re.search(r"Email:\s*(\S+@\S+)", text)
        if email_match:
            data["email"] = email_match.group(1).strip()

    # Phone
    phone_match = re.search(r"Phone:\s*([\d\-\(\)\.\s]+)", text)
    if phone_match:
        data["phone"] = phone_match.group(1).strip()

    # Website — look for http links that aren't locate.church
    for a in soup.find_all("a", href=re.compile(r"^https?://")):
        href = a["href"]
        if "locate.church" not in href and "thegospelcoalition" not in href:
            data["website"] = href
            break

    # Address — the Contact Information section
    # Look for street address patterns
    addr_match = re.search(r"(\d+\s+[\w\s\.]+(?:Street|St|Avenue|Ave|Road|Rd|Drive|Dr|Boulevard|Blvd|Lane|Ln|Way|Circle|Ct|Court|Pike|Highway|Hwy)[\w\s\.]*)", text, re.IGNORECASE)
    if addr_match:
        data["address"] = addr_match.group(1).strip()

    # Networks (detail page may have more accurate info)
    networks = []
    net_match = re.search(r"Networks?:\s*(.+?)(?:\n|$)", text)
    if net_match:
        raw = net_match.group(1).strip()
        networks = [n.strip() for n in raw.split(",") if n.strip()]
    data["networks_detail"] = networks

    return data


# ============================================================
# PROGRESS MANAGEMENT
# ============================================================
def save_progress(churches, completed_networks, current_network, current_page):
    with open(PROGRESS_FILE, "w") as f:
        json.dump({
            "churches": churches,
            "completed_networks": completed_networks,
            "current_network": current_network,
            "current_page": current_page,
            "timestamp": datetime.now().isoformat(),
        }, f)


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return None


# ============================================================
# MAIN SCRAPE LOGIC
# ============================================================
def scrape_network(network_name, network_path, max_pages=None):
    """Scrape all pages of a single network. Returns list of church dicts."""
    churches = []
    page = 1
    total_expected = None

    # Try primary URL first
    base_list_url = f"{BASE_URL}{network_path}"
    url = base_list_url

    log.info(f"Starting network: {network_name} — {url}")

    while True:
        if max_pages and page > max_pages:
            log.info(f"Reached max pages ({max_pages}) for {network_name}")
            break

        log.info(f"  Fetching page {page}..." + (f" ({len(churches)} churches so far)" if churches else ""))
        html = fetch_page(url)

        if html is None:
            # Try fallback URLs on first page
            if page == 1 and network_name in NETWORK_FALLBACKS:
                for fallback_path in NETWORK_FALLBACKS[network_name]:
                    fallback_url = f"{BASE_URL}{fallback_path}"
                    log.info(f"  Trying fallback: {fallback_url}")
                    html = fetch_page(fallback_url)
                    if html:
                        base_list_url = fallback_url
                        break
            if html is None:
                log.error(f"  Could not fetch page {page}, stopping network {network_name}")
                break

        page_churches, next_url, total = parse_list_page(html, network_name)

        if total and not total_expected:
            total_expected = total
            log.info(f"  Total churches in {network_name}: {total}")

        if not page_churches:
            log.info(f"  No churches found on page {page}, done with {network_name}")
            break

        churches.extend(page_churches)
        log.info(f"  Page {page}: {len(page_churches)} churches (running total: {len(churches)})")

        # Determine next page URL
        if next_url:
            url = next_url
        else:
            # Try incrementing page parameter
            next_page_url = f"{base_list_url}?page={page + 1}"
            url = next_page_url

        page += 1

        # Safety valve
        if page > 500:
            log.warning("Safety valve: exceeded 500 pages, stopping")
            break

    log.info(f"Finished {network_name}: {len(churches)} churches collected")
    return churches


def deduplicate(all_churches):
    """Merge churches that appear in multiple networks by slug."""
    by_slug = {}
    for church in all_churches:
        slug = church["slug"]
        if slug in by_slug:
            # Merge networks
            existing = by_slug[slug]
            for net in church["networks"]:
                if net not in existing["networks"]:
                    existing["networks"].append(net)
            # Keep longer description
            if len(church.get("description", "")) > len(existing.get("description", "")):
                existing["description"] = church["description"]
        else:
            by_slug[slug] = church.copy()
    return list(by_slug.values())


def enrich_with_details(churches, skip_details=False):
    """Visit each church's detail page to get pastor/email/phone/website."""
    if skip_details:
        log.info("Skipping detail enrichment (--skip-details)")
        return churches

    total = len(churches)
    log.info(f"Enriching {total} churches with detail page data...")

    for i, church in enumerate(churches):
        if i % 50 == 0 and i > 0:
            log.info(f"  Progress: {i}/{total} ({i*100//total}%)")

        url = church.get("detail_url")
        if not url:
            continue

        html = fetch_detail(url)
        if not html:
            continue

        detail = parse_detail_page(html)
        church["pastor"] = detail.get("pastor", "")
        church["email"] = detail.get("email", "")
        church["phone"] = detail.get("phone", "")
        church["address"] = detail.get("address", "")
        church["website"] = detail.get("website", "")

        # Update networks if detail page has better data
        if detail.get("networks_detail"):
            for net in detail["networks_detail"]:
                if net not in church["networks"]:
                    church["networks"].append(net)

    log.info(f"Enrichment complete")
    return churches


# ============================================================
# OUTPUT — CSV (intermediate)
# ============================================================
def write_csv(churches, filepath):
    fieldnames = [
        "name", "slug", "pastor", "email", "phone", "website",
        "location_raw", "city", "state", "zip", "address",
        "networks", "is_9marks", "is_founders", "is_tgc",
        "description", "detail_url",
    ]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for church in churches:
            city, state, zipcode = parse_location(church.get("location_raw", ""))
            nets = church.get("networks", [])
            nets_lower = [n.lower() for n in nets]
            writer.writerow({
                "name": church.get("name", ""),
                "slug": church.get("slug", ""),
                "pastor": church.get("pastor", ""),
                "email": church.get("email", ""),
                "phone": church.get("phone", ""),
                "website": church.get("website", ""),
                "location_raw": church.get("location_raw", ""),
                "city": city,
                "state": state,
                "zip": zipcode,
                "address": church.get("address", ""),
                "networks": ", ".join(nets),
                "is_9marks": "Yes" if any("9marks" in n or "9 marks" in n for n in nets_lower) else "No",
                "is_founders": "Yes" if any("founders" in n for n in nets_lower) else "No",
                "is_tgc": "Yes" if any("gospel coalition" in n or "tgc" in n for n in nets_lower) else "No",
                "description": church.get("description", ""),
                "detail_url": church.get("detail_url", ""),
            })
    log.info(f"CSV written: {filepath} ({len(churches)} rows)")


# ============================================================
# OUTPUT — XLSX (CRM format)
# ============================================================
def write_xlsx(churches, filepath):
    """Write to the SG Church CRM format spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="Arial", size=10, color="333333")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Church Leads"
    ws.sheet_properties.tabColor = "1A1A2E"

    headers = [
        "Lead ID", "Church Name", "Senior Pastor", "Pastor Email",
        "Church Email", "Phone", "Website", "City", "State", "ZIP",
        "Denomination", "Est. Congregation Size", "Directory Source(s)",
        "9Marks Listed", "Founders Listed", "TGC Listed",
        "Fiscal Year Model", "Sermon Archive Online?", "Archive Platform",
        "Est. Sermon Count", "Lead Status", "Lead Score",
        "First Contact Date", "Last Contact Date", "Contact Method",
        "Notes", "Next Action", "Next Action Date", "Assigned To",
        "Source URL",
    ]

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    widths = [10, 30, 22, 28, 28, 16, 32, 16, 8, 10, 20, 14, 22,
              12, 12, 10, 16, 14, 18, 14, 16, 10, 14, 14, 14, 32, 24, 14, 14, 36]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Write church data
    for idx, church in enumerate(churches):
        row = idx + 2
        lead_id = f"SG-{idx+1:04d}"

        city, state, zipcode = parse_location(church.get("location_raw", ""))
        nets = church.get("networks", [])
        nets_lower = [n.lower() for n in nets]

        is_9m = "Yes" if any("9marks" in n or "9 marks" in n for n in nets_lower) else "No"
        is_fn = "Yes" if any("founders" in n for n in nets_lower) else "No"
        is_tgc = "Yes" if any("gospel coalition" in n or "tgc" in n for n in nets_lower) else "No"

        # Calculate lead score
        score = 0
        directory_count = sum(1 for x in [is_9m, is_fn, is_tgc] if x == "Yes")
        if directory_count >= 2:
            score += 3
        elif directory_count == 1:
            score += 1

        row_data = [
            lead_id,
            church.get("name", ""),
            church.get("pastor", ""),
            church.get("email", ""),
            "",  # Church Email (separate from pastor email)
            church.get("phone", ""),
            church.get("website", ""),
            city,
            state,
            zipcode,
            "",  # Denomination (to be enriched)
            "",  # Est. Congregation Size
            ", ".join(nets),
            is_9m,
            is_fn,
            is_tgc,
            "",  # Fiscal Year Model
            "",  # Sermon Archive Online?
            "",  # Archive Platform
            "",  # Est. Sermon Count
            "New",  # Lead Status
            score if score > 0 else "",
            "",  # First Contact Date
            "",  # Last Contact Date
            "",  # Contact Method
            church.get("description", ""),
            "",  # Next Action
            "",  # Next Action Date
            "",  # Assigned To
            church.get("detail_url", ""),
        ]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # Data validations (only if we have data rows)
    max_row = max(len(churches) + 1, 2)
    dv_status = DataValidation(type="list", formula1='"New,Research,Contacted,Warm,Qualified,Proposal Sent,Won,Lost,Nurture"')
    ws.add_data_validation(dv_status)
    dv_status.add(f"U2:U{max_row}")

    dv_fiscal = DataValidation(type="list", formula1='"Calendar (Jan-Dec),Ministry (Sep-Aug),Unknown,Other"')
    ws.add_data_validation(dv_fiscal)
    dv_fiscal.add(f"Q2:Q{max_row}")

    for col_range in ["N", "O", "P", "R"]:
        dv = DataValidation(type="list", formula1='"Yes,No,Unknown"')
        ws.add_data_validation(dv)
        dv.add(f"{col_range}2:{col_range}{max_row}")

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Scrape Summary")
    ws2.sheet_properties.tabColor = "C5A55A"
    summary = [
        ["Shepherd's Guild — Locate.Church Scrape Results"],
        [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        [""],
        ["Metric", "Value"],
        ["Total churches (deduplicated)", len(churches)],
        ["Churches on 9Marks", sum(1 for c in churches if any("9marks" in n.lower() or "9 marks" in n.lower() for n in c.get("networks", [])))],
        ["Churches on Founders", sum(1 for c in churches if any("founders" in n.lower() for n in c.get("networks", [])))],
        ["Churches on TGC", sum(1 for c in churches if any("gospel coalition" in n.lower() or "tgc" in n.lower() for n in c.get("networks", [])))],
        ["Multi-directory (2+)", sum(1 for c in churches if sum(1 for check in [
            any("9marks" in n.lower() or "9 marks" in n.lower() for n in c.get("networks", [])),
            any("founders" in n.lower() for n in c.get("networks", [])),
            any("gospel coalition" in n.lower() or "tgc" in n.lower() for n in c.get("networks", [])),
        ] if check) >= 2)],
        ["With pastor name", sum(1 for c in churches if c.get("pastor"))],
        ["With email", sum(1 for c in churches if c.get("email"))],
        ["With phone", sum(1 for c in churches if c.get("phone"))],
        ["With website", sum(1 for c in churches if c.get("website"))],
    ]
    for r, row_data in enumerate(summary, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1A1A2E")
    ws2.merge_cells("A1:B1")
    ws2["A4"].font = Font(name="Arial", bold=True, size=10)
    ws2["B4"].font = Font(name="Arial", bold=True, size=10)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    wb.save(filepath)
    log.info(f"XLSX written: {filepath} ({len(churches)} rows)")


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Scrape Locate.Church directories for SG CRM")
    parser.add_argument("--max-pages", type=int, default=None, help="Max pages per network (default: all)")
    parser.add_argument("--skip-details", action="store_true", help="Skip detail page enrichment")
    parser.add_argument("--networks", nargs="+", choices=["9marks", "founders", "tgc"], default=["9marks", "founders", "tgc"])
    parser.add_argument("--resume", action="store_true", help="Resume from progress file")
    parser.add_argument("--output-dir", default=".", help="Output directory")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    all_churches = []
    completed_networks = []

    # Resume support
    if args.resume:
        progress = load_progress()
        if progress:
            all_churches = progress.get("churches", [])
            completed_networks = progress.get("completed_networks", [])
            log.info(f"Resumed with {len(all_churches)} churches from {len(completed_networks)} networks")

    # Scrape each network
    for net_name in args.networks:
        if net_name in completed_networks:
            log.info(f"Skipping {net_name} (already completed)")
            continue

        net_path = NETWORKS[net_name]
        churches = scrape_network(net_name, net_path, max_pages=args.max_pages)
        all_churches.extend(churches)
        completed_networks.append(net_name)

        # Save progress after each network
        save_progress(all_churches, completed_networks, net_name, -1)

    # Deduplicate
    log.info(f"Total raw: {len(all_churches)} churches")
    deduped = deduplicate(all_churches)
    log.info(f"After dedup: {len(deduped)} unique churches")

    # Enrich with detail pages
    deduped = enrich_with_details(deduped, skip_details=args.skip_details)

    # Write outputs
    csv_path = str(output_dir / OUTPUT_CSV)
    xlsx_path = str(output_dir / OUTPUT_XLSX)

    write_csv(deduped, csv_path)
    write_xlsx(deduped, xlsx_path)

    # Print summary
    print("\n" + "=" * 60)
    print("SCRAPE COMPLETE")
    print("=" * 60)
    print(f"Total unique churches: {len(deduped)}")

    nets_9m = sum(1 for c in deduped if any("9marks" in n.lower() or "9 marks" in n.lower() for n in c.get("networks", [])))
    nets_fn = sum(1 for c in deduped if any("founders" in n.lower() for n in c.get("networks", [])))
    nets_tgc = sum(1 for c in deduped if any("gospel coalition" in n.lower() or "tgc" in n.lower() for n in c.get("networks", [])))
    multi = sum(1 for c in deduped if sum(1 for check in [
        any("9marks" in n.lower() or "9 marks" in n.lower() for n in c.get("networks", [])),
        any("founders" in n.lower() for n in c.get("networks", [])),
        any("gospel coalition" in n.lower() or "tgc" in n.lower() for n in c.get("networks", [])),
    ] if check) >= 2)

    print(f"  9Marks: {nets_9m}")
    print(f"  Founders: {nets_fn}")
    print(f"  TGC: {nets_tgc}")
    print(f"  Multi-directory (2+): {multi}")
    with_pastor = sum(1 for c in deduped if c.get("pastor"))
    with_email = sum(1 for c in deduped if c.get("email"))
    print(f"  With pastor name: {with_pastor}")
    print(f"  With email: {with_email}")
    print(f"\nOutputs:")
    print(f"  CSV: {csv_path}")
    print(f"  XLSX: {xlsx_path}")
    print("=" * 60)

    # Cleanup progress file
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)


if __name__ == "__main__":
    main()
